"""
Convert the full EFCAMDAT Excel (1.2M rows) to data/raw/efcamdat_full.csv.
Streams rows so memory stays low. Run from project root:

  python xlsx_to_csv.py [path_to_excel.xlsx]

If no path given, uses: data/raw/Final database (alternative prompts).xlsx
"""

import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Install openpyxl: pip install openpyxl")

PROJECT_ROOT = Path(__file__).resolve().parent
OUTPUT_CSV = PROJECT_ROOT / "data" / "raw" / "efcamdat_full.csv"
WANTED = ("l1", "cefr", "text", "cefr_numeric", "wordcount", "topic", "grade", "nationality")


def main() -> None:
    if len(sys.argv) > 1:
        xlsx_path = Path(sys.argv[1])
    else:
        xlsx_path = PROJECT_ROOT / "data" / "raw" / "Final database (alternative prompts).xlsx"
    if not xlsx_path.exists():
        print(f"Not found: {xlsx_path}")
        print("Usage: python xlsx_to_csv.py [path_to_excel.xlsx]")
        sys.exit(1)

    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    # First row = headers
    header = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_index = {h: i for i, h in enumerate(header) if h}
    missing = [c for c in WANTED if c not in col_index]
    if missing:
        print(f"Missing columns in Excel: {missing}. Found: {list(col_index.keys())}")
        wb.close()
        sys.exit(1)

    written = 0
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(WANTED)
        for row in ws.iter_rows(min_row=2):
            values = [row[i].value for i in range(len(header)) if i < len(row)]
            if len(values) <= max(col_index.values()):
                continue
            out_row = []
            for col in WANTED:
                val = values[col_index[col]] if col_index[col] < len(values) else None
                out_row.append("" if val is None else str(val).strip())
            writer.writerow(out_row)
            written += 1
            if written % 100_000 == 0:
                print(f"  {written} rows...")
    wb.close()
    print(f"Wrote {written} rows to {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
